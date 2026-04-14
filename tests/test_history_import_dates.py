"""送货历史导入：日期解析（-、/、Excel 序列日）与模板。"""

import asyncio
from datetime import date, datetime, timezone
from io import BytesIO
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import load_workbook

import app.intelligent_prediction.services.history_service as history_service_mod
from app.intelligent_prediction.services.history_service import HistoryService


@pytest.fixture
def svc() -> HistoryService:
    return HistoryService()


def test_parse_iso_hyphen(svc: HistoryService) -> None:
    d, err = svc._parse_date_cell("2026-01-05")
    assert err is None
    assert d == date(2026, 1, 5)


def test_parse_iso_slash(svc: HistoryService) -> None:
    d, err = svc._parse_date_cell("2026/1/5")
    assert err is None
    assert d == date(2026, 1, 5)


def test_parse_cn_full_year_month_day(svc: HistoryService) -> None:
    d, err = svc._parse_date_cell("2026年1月9日")
    assert err is None
    assert d == date(2026, 1, 9)


def test_parse_cn_month_day_uses_utc_year(svc: HistoryService) -> None:
    fixed_now = datetime(2026, 6, 1, 12, 0, 0, tzinfo=timezone.utc)
    with patch.object(history_service_mod.datetime, "now", return_value=fixed_now):
        d, err = svc._parse_date_cell("1月9日")
    assert err is None
    assert d == date(2026, 1, 9)


def test_parse_cn_month_day_hao_variant(svc: HistoryService) -> None:
    fixed_now = datetime(2026, 6, 1, 12, 0, 0, tzinfo=timezone.utc)
    with patch.object(history_service_mod.datetime, "now", return_value=fixed_now):
        d, err = svc._parse_date_cell("12月31号")
    assert err is None
    assert d == date(2026, 12, 31)


def test_parse_with_time(svc: HistoryService) -> None:
    d, err = svc._parse_date_cell("2026/03/15 00:00:00")
    assert err is None
    assert d == date(2026, 3, 15)


def test_parse_excel_serial(svc: HistoryService) -> None:
    d, err = svc._parse_date_cell(46027)
    assert err is None
    assert d == date(2026, 1, 5)


def test_parse_datetime_pandas(svc: HistoryService) -> None:
    d, err = svc._parse_date_cell(pd.Timestamp("2025-12-01"))
    assert err is None
    assert d == date(2025, 12, 1)


def test_parse_datetime_native(svc: HistoryService) -> None:
    d, err = svc._parse_date_cell(datetime(2024, 6, 30, 12, 0, 0))
    assert err is None
    assert d == date(2024, 6, 30)


def test_import_template_xlsx_has_data_and_help_sheets() -> None:
    raw = HistoryService.import_template_xlsx_bytes()
    wb = load_workbook(BytesIO(raw), read_only=True)
    assert wb.sheetnames[0] == "导入数据"
    assert "使用说明" in wb.sheetnames
    ws = wb["导入数据"]
    assert list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))) == HistoryService.import_template_headers()
    row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    assert str(row2[0] or "").startswith(HistoryService.TEMPLATE_EXAMPLE_RM_PREFIX)


def test_import_skips_example_rows_in_official_template() -> None:
    class _Sess:
        def __init__(self) -> None:
            self.added: list[object] = []

        def add(self, obj: object) -> None:
            self.added.append(obj)

        async def commit(self) -> None:
            return None

        async def rollback(self) -> None:
            return None

    async def _run() -> None:
        svc = HistoryService()
        raw = HistoryService.import_template_xlsx_bytes()
        sess = _Sess()
        res = await svc.import_excel(sess, raw, "template.xlsx")
        assert res.inserted == 0
        assert res.skipped == 3
        assert res.errors == []
        assert len(sess.added) == 0

    asyncio.run(_run())
